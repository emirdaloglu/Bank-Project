# =============================================================================
# BANKA MÜŞTERİ ŞİKAYETLERİ SINIFLANDIRMA SCRIPTİ
# =============================================================================
# 
# Bu Python scripti, banka müşteri şikayetlerini otomatik olarak kategorilere 
# ayırmak için gelişmiş makine öğrenmesi modelleri eğitir ve kullanır.
# 
# AMAÇ:
# - Excel dosyasından müşteri şikayetlerini okuma
# - Etiketli verilerle model eğitimi
# - Etiketsiz verilerin kategorilerini tahmin etme
# - Model performansını değerlendirme ve görselleştirme
# - Sonuçları Excel dosyasına kaydetme
# 
# ÖZELLİKLER:
# - TF-IDF, SBERT ve BERT modellerini destekler
# - Ensemble yaklaşımı (word + char TF-IDF)
# - Kural tabanlı override sistemi
# - Pseudo-labeling desteği
# - Otomatik hiperparametre optimizasyonu
# - Görselleştirme ve raporlama
# 
# KULLANIM:
# python svm_fill_categories.py
# 
# GİRDİ: musteri sikayetleri.xlsx
# ÇIKTI: musteri_sikayetleri_tahminli.xlsx + görselleştirmeler
# =============================================================================

import os
import sys
import json
import re
import warnings
from typing import Tuple, Optional

import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split, GridSearchCV, StratifiedKFold
from sklearn.pipeline import Pipeline
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix
from sklearn.svm import LinearSVC
from sklearn.base import BaseEstimator, TransformerMixin
import torch
from torch.utils.data import Dataset
from transformers import (
    AutoTokenizer,
    AutoModelForSequenceClassification,
    DataCollatorWithPadding,
)
from torch.utils.data import DataLoader
from torch.optim import AdamW
import matplotlib.pyplot as plt
import seaborn as sns
from wordcloud import WordCloud
from openpyxl import load_workbook
try:
    from sentence_transformers import SentenceTransformer
    HAS_SENTENCE_TRANSFORMERS = True
except Exception:
    HAS_SENTENCE_TRANSFORMERS = False

# =============================================================================
# TÜRKÇE STOPWORDS LİSTESİ
# =============================================================================
# Bu liste, metin analizi sırasında göz ardı edilecek Türkçe kelimeleri içerir
# Bu kelimeler genellikle anlam taşımaz ve sınıflandırma performansını düşürür
# 
# İçerik:
# - Bağlaçlar: ve, veya, ile, ama, fakat, lakin
# - Zamirler: bu, şu, o, ben, sen, biz, siz, onlar
# - Edatlar: için, gibi, üzerine, hakkında
# - Soru kelimeleri: ne, neden, nasıl, niçin
# - Zaman belirteçleri: önce, sonra, artık
# - Diğer gereksiz kelimeler
# =============================================================================
TURKISH_STOPWORDS = set(
    [
        "ve","veya","ile","de","da","ki","bu","şu","o","bir","biraz","çok","az","daha",
        "en","gibi","için","çünkü","ama","fakat","lakin","ile","mi","mu","mü","mı","ne",
        "neden","nasıl","niçin","niye","beni","bana","sen","senin","siz","sizin","biz","bizim",
        "onlar","onların","ben","var","yok","hem","her","herhangi","hepsi","bazı","diğer",
        "şey","şeyi","şeyler","olan","olarak","üzerine","üzerinden","hakkında","daha","sonra",
        "önce","mı","mi","mu","mü","ise","yine","artık","ancak","yada","ya","bile","kadar",
        "hep","hiç","hiçbir","kez","yani","zaten","olan","oldu","oluyor","olacak","ettim","etti",
        "ediyor","edecek","gün","ay","yıl","vs","vb","v.s","v.b",
    ]
)

# Output/plot controls
GENERATE_DIAGNOSTIC_PLOTS = True
GENERATE_OVERALL_WORDCLOUD = True
GENERATE_PER_CLASS_WORDCLOUDS = False

# Vectorization mode: 'tfidf' or 'sbert'
# Defaulting to 'tfidf' as it performed better on this dataset
VECTORIZER_MODE = 'tfidf'

# Advanced options
ENSEMBLE_TFIDF = True  # Use word + char TF-IDF ensemble
ENABLE_PSEUDO_LABELING = False
PSEUDO_LABEL_THRESHOLD = 0.85
PREFER_CHAR_ONLY = False  # prefer ensemble as it performed better

# =============================================================================
# KURAL TABANLI OVERRIDE SİSTEMİ
# =============================================================================
# Bu sistem, model tahminlerini belirli anahtar kelimelere göre düzeltir
# Özellikle düşük güvenilirlikli tahminlerde kullanışlıdır
# 
# AYARLAR:
# - ENABLE_RULE_OVERRIDES: Kural sistemini aktif/pasif yapar
# - RULE_OVERRIDE_CONF_THRESHOLD: Bu eşiğin altındaki güvenilirlikte kurallar uygulanır
# - RULE_MIN_MATCHES_STRONG: Güçlü override için minimum anahtar kelime eşleşmesi
# - RULE_MIN_MATCHES_WEAK: Zayıf override için minimum anahtar kelime eşleşmesi
# =============================================================================
ENABLE_RULE_OVERRIDES = False
RULE_OVERRIDE_CONF_THRESHOLD = 0.75  # apply rules if model confidence is below this
RULE_MIN_MATCHES_STRONG = 2          # strong override if >= 2 keyword hits
RULE_MIN_MATCHES_WEAK = 1            # weak override if >= 1 and low confidence

# =============================================================================
# KATEGORİ BAZLI ANAHTAR KELİMELER
# =============================================================================
# Her kategori için o kategoriyi belirleyen anahtar kelimeler
# Bu kelimeler metinde bulunduğunda, o kategoriye atama yapılır
# =============================================================================
RULE_KEYWORDS = {
    "Kur Marjı": [
        "kur","marj","makas","spread","altın","gram","bozdum","döviz","doviz","alış","satiş","satış","fiyat","piyasa","ons"
    ],
    "Çalışmıyor - Hata Veriyor": [
        "çalışmıyor","calismiyor","hata","açılmıyor","acilmiyor","çöktü","coktu","vermez","donuyor","takılıyor","baglanmiyor","bağlanmıyor","açılmıyor","hata veriyor"
    ],
    "Kart": [
        "kart","kredi kart","limit","ekstre","pos","temassız","temassiz","cvv","mail order","temdit"
    ],
    "Şube": [
        "şube","sube","gişe","gise","personel","müdür","mudur","vezne","sıra","sira"
    ],
    "Yavaş": [
        "yavaş","yavas","gecikme","yavaşladı","lag","ağır","agir","yavaşlama"
    ],
    "Çağrı Merkezi": [
        "çağrı merkezi","cagri merkezi","müşteri hizmetleri","musteri hizmetleri",
        "telefonla ulaş","telefonla ulas","ulaşamıyorum","ulasamiyorum","ulaşmak", "ulaşamıyoruz","cagriya ulasamiyorum",
        "ulaşma süresi","ulasma suresi","hat meşgul","hat mesgul","bekletiyor","bekleme süresi","bekleme suresi"
    ],
    "Kullanışsız": [
        "kullanışsız","kullanisiz","kullanımı zor","kullanimi zor","menü","menu","arayüz","arayuz","ui","ux"
    ],
    "AlbaFX": ["alba","fx","alba fx"],
    "Aşırı Güvenlik": ["güvenlik","guvenlik","otp","sms","şifre","sifre","doğrulama","dogrulama","captcha"]
}

def rule_override_category(text: str, current_label: str, confidence: float) -> tuple[str, str]:
    """
    Kural tabanlı kategori override sistemi
    
    Bu fonksiyon, model tahminini belirli anahtar kelimelere göre düzeltir.
    Özellikle düşük güvenilirlikli tahminlerde kullanışlıdır.
    
    Parametreler:
    text (str): Analiz edilecek metin
    current_label (str): Modelin mevcut tahmini
    confidence (float): Model tahmin güvenilirliği (0-1 arası)
    
    Döndürür:
    tuple[str, str]: (düzeltilmiş_etiket, düzeltme_nedeni)
    
    Mantık:
    1. Metinde her kategori için anahtar kelime sayısını hesaplar
    2. En çok anahtar kelime eşleşen kategoriyi bulur
    3. Güçlü override: >= 2 anahtar kelime varsa kesin değiştirir
    4. Zayıf override: Düşük güvenilirlik + >= 1 anahtar kelime varsa değiştirir
    """
    t = str(text).lower()  # Metni küçük harfe çevir
    
    # Her kategori için anahtar kelime eşleşme sayısını hesapla
    best_label = None
    best_hits = 0
    for label, keywords in RULE_KEYWORDS.items():
        hits = 0
        for kw in keywords:
            if kw in t:  # Anahtar kelime metinde var mı?
                hits += 1
        if hits > best_hits:  # En çok eşleşen kategoriyi bul
            best_hits = hits
            best_label = label
    
    # Override kararı ver
    reason = ""
    
    # Güçlü override: Yeterli anahtar kelime varsa kesin değiştir
    if best_hits >= RULE_MIN_MATCHES_STRONG and best_label and best_label != current_label:
        reason = f"rule_strong({best_label}:{best_hits})"
        return best_label, reason
    
    # Zayıf override: Düşük güvenilirlik + bazı anahtar kelimeler varsa değiştir
    if confidence < RULE_OVERRIDE_CONF_THRESHOLD and best_hits >= RULE_MIN_MATCHES_WEAK and best_label and best_label != current_label:
        reason = f"rule_weak({best_label}:{best_hits},conf={confidence:.2f})"
        return best_label, reason
    
    # Değişiklik yok, mevcut etiketi koru
    return current_label, reason

# Suggestions for all rows (do not modify original labels)
ADD_SUGGESTIONS_FOR_ALL = False
SBERT_MODEL_NAME = 'sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2'


def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame sütun isimlerini normalize eder
    
    Bu fonksiyon, Excel dosyasından okunan sütun isimlerini standartlaştırır:
    - Tüm harfleri küçük harfe çevirir
    - Başındaki ve sonundaki boşlukları kaldırır
    - String formatına çevirir
    
    Parametreler:
    df (pd.DataFrame): Normalize edilecek DataFrame
    
    Döndürür:
    pd.DataFrame: Sütun isimleri normalize edilmiş DataFrame
    """
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def find_required_columns(df: pd.DataFrame) -> Tuple[str, str, str]:
    """
    DataFrame'de gerekli sütunları bulur ve döndürür
    
    Bu fonksiyon, Excel dosyasında tarih, yorum ve kategori sütunlarını
    farklı isim varyasyonlarıyla bulur. Büyük/küçük harf ve boşluk duyarsızdır.
    
    Parametreler:
    df (pd.DataFrame): Sütunları aranacak DataFrame
    
    Döndürür:
    Tuple[str, str, str]: (tarih_sütunu, yorum_sütunu, kategori_sütunu)
    
    Hata:
    KeyError: Gerekli sütunlardan biri bulunamazsa
    """
    cols = {c: c for c in df.columns}
    
    # Türkçe sütun ismi varyasyonları
    tarih_aliases = {"tarih", "date", "olusturma_tarihi", "oluşturma_tarihi"}
    yorum_aliases = {"yorum", "text", "yorumlar", "şikayet", "sikayet", "aciklama", "açıklama"}
    kategori_aliases = {"kategori", "category", "etiket", "label"}

    def pick(alias_set: set) -> str:
        for alias in alias_set:
            if alias in cols:
                return cols[alias]
        # Fallback: try partial match
        for c in df.columns:
            base = c.replace("ı", "i").replace("İ", "i")
            for alias in alias_set:
                a = alias.replace("ı", "i").replace("İ", "i")
                if a in base:
                    return c
        raise KeyError(f"Kolon bulunamadı: {sorted(alias_set)}. Mevcut kolonlar: {list(df.columns)}")

    tarih_col = pick(tarih_aliases)
    yorum_col = pick(yorum_aliases)
    kategori_col = pick(kategori_aliases)
    return tarih_col, yorum_col, kategori_col


def load_excel_with_header_detection(input_path: str) -> pd.DataFrame:
    """
    Excel dosyasını yükler ve doğru başlık satırını otomatik tespit eder
    
    Bu fonksiyon, Excel dosyasında tarih/yorum/kategori sütunlarını içeren
    doğru sayfa ve başlık satırını otomatik olarak bulur. Farklı Excel
    formatlarını destekler ve esnek bir yükleme sağlar.
    
    Parametreler:
    input_path (str): Yüklenecek Excel dosyasının yolu
    
    Döndürür:
    pd.DataFrame: Yüklenen veri
    
    Özellikler:
    - Çoklu sayfa desteği
    - Otomatik başlık satırı tespiti
    - Türkçe sütun ismi varyasyonları desteği
    - Hata durumunda fallback mekanizması
    """
    xl = pd.ExcelFile(input_path, engine="openpyxl")

    # Türkçe sütun ismi varyasyonları
    tarih_aliases = {"tarih", "date", "olusturma_tarihi", "oluşturma_tarihi"}
    yorum_aliases = {"yorum", "text", "yorumlar", "şikayet", "sikayet", "aciklama", "açıklama", "comment"}
    kategori_aliases = {"kategori", "category", "etiket", "label"}

    def norm(s: object) -> str:
        t = str(s).strip().lower()
        t = t.replace("ı", "i").replace("İ", "i")
        return t

    best_choice: Optional[Tuple[str, int, int]] = None  # (sheet_name, header_row, score)
    for sheet in xl.sheet_names:
        df_raw = pd.read_excel(input_path, sheet_name=sheet, header=None, dtype=object, engine="openpyxl")
        max_rows_to_scan = min(50, len(df_raw))
        for row_idx in range(max_rows_to_scan):
            row_vals = [norm(v) for v in df_raw.iloc[row_idx].tolist()]
            # compute coverage
            def has_any(aliases: set) -> bool:
                for cell in row_vals:
                    for a in aliases:
                        a2 = norm(a)
                        if a2 and a2 in cell:
                            return True
                return False

            cov_tarih = has_any(tarih_aliases)
            cov_yorum = has_any(yorum_aliases)
            cov_kategori = has_any(kategori_aliases)
            score = int(cov_tarih) + int(cov_yorum) + int(cov_kategori)
            # Require en azindan yorum ve kategori'yi
            if cov_yorum and cov_kategori:
                if best_choice is None or score > best_choice[2]:
                    best_choice = (sheet, row_idx, score)

    if best_choice is not None:
        sheet, header_row, _ = best_choice
        df = pd.read_excel(input_path, sheet_name=sheet, header=header_row, engine="openpyxl")
        return df

    # Fallback
    return pd.read_excel(input_path, sheet_name=xl.sheet_names[0], header=0, engine="openpyxl")


def build_pipeline() -> Pipeline:
    """
    TF-IDF + LinearSVC sınıflandırma pipeline'ı oluşturur
    
    Bu fonksiyon, metin sınıflandırma için standart bir pipeline oluşturur:
    1. TF-IDF vektörizasyonu: Metinleri sayısal özelliklere dönüştürür
    2. LinearSVC sınıflandırıcı: Destek vektör makinesi ile sınıflandırma
    
    Pipeline Özellikleri:
    - lowercase=True: Tüm metinleri küçük harfe çevirir
    - ngram_range=(1,2): Tek kelime ve kelime çiftlerini kullanır
    - min_df=2: En az 2 dokümanda geçen kelimeleri alır
    - max_df=0.9: %90'dan fazla dokümanda geçen kelimeleri çıkarır
    - sublinear_tf=True: TF-IDF hesaplamasında logaritmik ölçekleme
    - max_features=100000: Maksimum özellik sayısı
    - stop_words: Türkçe stopwords'leri kullanır
    - strip_accents: Aksanları kaldırır
    
    Döndürür:
    Pipeline: Eğitilmeye hazır sklearn pipeline'ı
    """
    pipeline = Pipeline(
        steps=[
            (
                "tfidf",
                TfidfVectorizer(
                    lowercase=True,
                    ngram_range=(1, 2),
                    min_df=2,
                    max_df=0.9,
                    sublinear_tf=True,
                    max_features=100000,
                    stop_words=sorted(list(TURKISH_STOPWORDS)),
                    strip_accents="unicode",
                ),
            ),
            ("clf", LinearSVC(class_weight="balanced", random_state=42)),
        ]
    )
    return pipeline


def build_sbert_encoder() -> Optional['SentenceTransformer']:
    if not HAS_SENTENCE_TRANSFORMERS:
        return None
    model = SentenceTransformer(SBERT_MODEL_NAME)
    return model


def tune_hyperparameters(X: pd.Series, y: pd.Series) -> Pipeline:
    """
    Hiperparametre optimizasyonu ile model performansını iyileştirir
    
    Bu fonksiyon, GridSearchCV kullanarak en iyi hiperparametreleri bulur.
    Farklı vektörizer modları için farklı optimizasyon stratejileri kullanır.
    
    Parametreler:
    X (pd.Series): Eğitim metinleri
    y (pd.Series): Hedef kategoriler
    
    Döndürür:
    Pipeline: En iyi hiperparametrelerle eğitilmiş model
    
    Optimizasyon Stratejileri:
    - TF-IDF: n-gram, min_df, max_df, C parametreleri
    - SBERT: Sadece C parametresi
    - BERT: Fine-tuning ile eğitim
    """
    if VECTORIZER_MODE == 'tfidf':
        pipeline = build_pipeline()
        param_grid = {
            # Try word/char analyzers with varied n-grams and limits
            "tfidf__analyzer": ["word", "char", "char_wb"],
            "tfidf__ngram_range": [(1, 1), (1, 2), (1, 3), (2, 3), (3, 5)],
            "tfidf__min_df": [1, 2],
            "tfidf__max_df": [0.9, 1.0],
            "tfidf__max_features": [100000, 200000],
            "clf__C": [0.25, 0.5, 1.0, 2.0, 4.0],
        }
        # Ensure n_splits does not exceed the smallest class count
        min_class_count = int(y.value_counts().min()) if y.nunique() > 1 else 2
        n_splits = max(2, min(5, min_class_count))
        cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
        grid = GridSearchCV(
            estimator=pipeline,
            param_grid=param_grid,
            scoring="accuracy",
            cv=cv,
            n_jobs=-1,
            verbose=0,
        )
        grid.fit(X, y)
        return grid.best_estimator_
    elif VECTORIZER_MODE == 'sbert':
        class SBERTVectorizer(BaseEstimator, TransformerMixin):
            def __init__(self, model_name: str = SBERT_MODEL_NAME, batch_size: int = 64):
                self.model_name = model_name
                self.batch_size = batch_size
                self.model = None

            def fit(self, X, y=None):
                if self.model is None:
                    self.model = build_sbert_encoder()
                return self

            def transform(self, X):
                if self.model is None:
                    self.model = build_sbert_encoder()
                embeddings = self.model.encode(list(X), batch_size=self.batch_size, show_progress_bar=False, normalize_embeddings=True)
                return embeddings

        pipeline = Pipeline(steps=[("sbert", SBERTVectorizer()), ("clf", LinearSVC(class_weight="balanced", random_state=42))])
        param_grid = {"clf__C": [0.25, 0.5, 1.0, 2.0, 4.0]}
        min_class_count = int(y.value_counts().min()) if y.nunique() > 1 else 2
        n_splits = max(2, min(5, min_class_count))
        cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
        grid = GridSearchCV(
            estimator=pipeline,
            param_grid=param_grid,
            scoring="accuracy",
            cv=cv,
            n_jobs=1,
            verbose=0,
        )
        grid.fit(X, y)
        return grid.best_estimator_
    else:
        # BERTurk fine-tune (compatible with MPS on Mac)
        BERT_MODEL_NAME = 'dbmdz/bert-base-turkish-cased'
        labels = sorted(y.unique())
        label2id = {l: i for i, l in enumerate(labels)}
        id2label = {i: l for l, i in label2id.items()}

        tokenizer = AutoTokenizer.from_pretrained(BERT_MODEL_NAME)

        class TextDataset(Dataset):
            def __init__(self, Xtext, ylabels=None):
                self.X = list(Xtext)
                self.y = None if ylabels is None else [label2id[v] for v in ylabels]
            def __len__(self):
                return len(self.X)
            def __getitem__(self, idx):
                item = tokenizer(self.X[idx], truncation=True, max_length=128)
                if self.y is not None:
                    item['labels'] = torch.tensor(self.y[idx], dtype=torch.long)
                return item

        # Split for quick tuning
        X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
        ds_train = TextDataset(X_train, y_train)
        ds_val = TextDataset(X_val, y_val)

        model = AutoModelForSequenceClassification.from_pretrained(BERT_MODEL_NAME, num_labels=len(labels), id2label=id2label, label2id=label2id)
        device = torch.device('mps' if torch.backends.mps.is_available() else ('cuda' if torch.cuda.is_available() else 'cpu'))
        model.to(device)

        data_collator = DataCollatorWithPadding(tokenizer=tokenizer)
        dl_train = DataLoader(ds_train, batch_size=8, shuffle=True, collate_fn=data_collator)
        dl_val = DataLoader(ds_val, batch_size=16, shuffle=False, collate_fn=data_collator)

        optimizer = AdamW(model.parameters(), lr=2e-5, weight_decay=0.01)
        model.train()
        epochs = 2
        for epoch in range(epochs):
            total_loss = 0.0
            for batch in dl_train:
                batch = {k: v.to(device) for k, v in batch.items()}
                optimizer.zero_grad()
                out = model(**batch)
                loss = out.loss
                loss.backward()
                torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
                optimizer.step()
                total_loss += float(loss.item())
            # simple val accuracy
            model.eval()
            correct = 0
            count = 0
            with torch.no_grad():
                for batch in dl_val:
                    labels_t = batch.pop('labels').to(device)
                    batch = {k: v.to(device) for k, v in batch.items()}
                    logits = model(**batch).logits
                    preds = logits.argmax(dim=-1)
                    correct += int((preds == labels_t).sum().item())
                    count += int(labels_t.numel())
            val_acc = correct / max(1, count)
            print(f"BERT epoch {epoch+1}: train_loss={total_loss/max(1,len(dl_train)):.4f} val_acc={val_acc:.4f}")
            model.train()

        class BertWrapper:
            def __init__(self, model, tokenizer, id2label):
                self.model = model
                self.tokenizer = tokenizer
                self.id2label = id2label
                self.model.eval()
                self.device = next(self.model.parameters()).device
            def fit(self, Xfull, yfull):
                return self
            def predict(self, Xtext):
                with torch.no_grad():
                    enc = self.tokenizer(list(Xtext), truncation=True, max_length=128, padding=True, return_tensors='pt')
                    enc = {k: v.to(self.device) for k, v in enc.items()}
                    logits = self.model(**enc).logits
                    preds = logits.argmax(dim=-1).cpu().numpy()
                    return np.array([self.id2label[int(i)] for i in preds])
            def decision_function(self, Xtext):
                with torch.no_grad():
                    enc = self.tokenizer(list(Xtext), truncation=True, max_length=128, padding=True, return_tensors='pt')
                    enc = {k: v.to(self.device) for k, v in enc.items()}
                    logits = self.model(**enc).logits
                    return logits.cpu().numpy()

        return BertWrapper(model, tokenizer, id2label)


def tune_tfidf_with_analyzer(X: pd.Series, y: pd.Series, analyzer: str) -> Pipeline:
    pipeline = build_pipeline()
    param_grid = {
        "tfidf__analyzer": [analyzer],
        "tfidf__ngram_range": ([(1, 1), (1, 2), (1, 3)] if analyzer == "word" else [(2, 4), (3, 5), (3, 6)]),
        "tfidf__min_df": [1, 2],
        "tfidf__max_df": [0.9, 1.0],
        "tfidf__max_features": [100000, 200000],
        "clf__C": [0.25, 0.5, 1.0, 2.0, 4.0],
    }
    min_class_count = int(y.value_counts().min()) if y.nunique() > 1 else 2
    n_splits = max(2, min(5, min_class_count))
    cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    grid = GridSearchCV(
        estimator=pipeline,
        param_grid=param_grid,
        scoring="accuracy",
        cv=cv,
        n_jobs=-1,
        verbose=0,
    )
    grid.fit(X, y)
    return grid.best_estimator_


def decision_scores(pipeline: Pipeline, X: pd.Series, common_classes: np.ndarray) -> np.ndarray:
    # Obtain decision scores and align to common class order
    scores = pipeline.decision_function(X)
    if scores.ndim == 1:
        scores = np.vstack([-scores, scores]).T
    clf = pipeline.named_steps.get("clf")
    model_classes = clf.classes_
    col_map = {c: i for i, c in enumerate(model_classes)}
    aligned = np.full((scores.shape[0], len(common_classes)), fill_value=-1e9, dtype=float)
    for j, c in enumerate(common_classes):
        if c in col_map:
            aligned[:, j] = scores[:, col_map[c]]
    return aligned


def predict_with_confidence_and_overrides(model, X_text: pd.Series, enable_overrides: bool = True):
    """
    Model tahminlerini güvenilirlik skorları ve override'larla birlikte yapar
    
    Bu fonksiyon, model tahminlerini gerçekleştirir ve güvenilirlik skorlarını hesaplar.
    Ayrıca kural tabanlı override sistemi ile tahminleri düzeltebilir.
    
    Parametreler:
    model: Eğitilmiş model (Pipeline, tuple veya custom model)
    X_text (pd.Series): Tahmin yapılacak metinler
    enable_overrides (bool): Kural tabanlı override sistemini aktif/pasif yapar
    
    Döndürür:
    tuple: (tahmin_etiketleri, güvenilirlik_skorları, override_nedenleri)
    
    Özellikler:
    - Ensemble model desteği (word + char TF-IDF)
    - Softmax ile güvenilirlik hesaplama
    - Kural tabanlı düzeltme sistemi
    - Detaylı override raporlama
    """
    # Returns predicted_labels (np.array[str]), max_proba (np.array[float]), overrides (list[str])
    if isinstance(model, tuple):
        final_word, final_char = model
        classes_union = np.unique(np.concatenate([final_word.named_steps['clf'].classes_, final_char.named_steps['clf'].classes_]))
        sw = decision_scores(final_word, X_text, classes_union)
        sc = decision_scores(final_char, X_text, classes_union)
        scores = sw + sc
        preds = classes_union[scores.argmax(axis=1)]
    else:
        preds = model.predict(X_text)
        scores = model.decision_function(X_text)
        if np.ndim(scores) == 1:
            scores = np.vstack([-scores, scores]).T
    # softmax for confidence
    try:
        scores = np.asarray(scores)
        scores = scores - scores.max(axis=1, keepdims=True)
        exp_scores = np.exp(scores)
        probs = exp_scores / exp_scores.sum(axis=1, keepdims=True)
        max_proba = probs.max(axis=1)
    except Exception:
        max_proba = np.full(shape=(len(X_text),), fill_value=np.nan)
    overrides = [""] * len(X_text)
    if enable_overrides and ENABLE_RULE_OVERRIDES:
        adjusted = []
        new_over = []
        for txt, lab, conf in zip(X_text.tolist(), preds, max_proba.tolist()):
            final_lab, reason = rule_override_category(txt, str(lab), float(conf) if conf == conf else 0.0)
            adjusted.append(final_lab)
            new_over.append(reason)
        preds = np.array(adjusted)
        overrides = new_over
    return np.array(preds), np.array(max_proba), overrides


# =============================================================================
# TOP-K TAHMİN ÇIKTILARI (Top-2/Top-3)
# =============================================================================
# Bu yardımcı fonksiyon, her kayıt için en olası ilk K sınıfı ve bunların
# olasılıklarını üretir. Ensemble (word+char TF-IDF) veya tek model fark etmeksizin
# decision_function skorlarını hizalar, toplar ve softmax ile olasılığa çevirir.
#
# Üretilen kolonlar (K=3 için):
# - tahmin_top1, tahmin_top1_conf
# - tahmin_top2, tahmin_top2_conf
# - tahmin_top3, tahmin_top3_conf
# =============================================================================
def predict_topk(model, X_text: pd.Series, k: int = 3):
    """
    Top-k tahminleri ve olasılıklarını döndürür.
    Ensemble (word+char) veya tek pipeline için decision_function skorlarını
    toplayıp softmax ile olasılığa çevirir.

    Döndürür:
    - topk_labels: np.ndarray[str], şekil (n, k)
    - topk_probs:  np.ndarray[float], şekil (n, k)
    - all_classes: np.ndarray[str], tüm sınıf isimleri (olasılık matrisinin kolon sırası)
    """
    # 1) Skor matrisi ve sınıf isimlerini hazırla
    if isinstance(model, tuple):
        final_word, final_char = model
        classes_union = np.unique(
            np.concatenate([
                final_word.named_steps['clf'].classes_,
                final_char.named_steps['clf'].classes_,
            ])
        )
        sw = decision_scores(final_word, X_text, classes_union)
        sc = decision_scores(final_char, X_text, classes_union)
        scores = sw + sc
        all_classes = classes_union
    else:
        scores = model.decision_function(X_text)
        if np.ndim(scores) == 1:
            # binary durumu -> iki kolonlu skor
            scores = np.vstack([-scores, scores]).T
        clf = model.named_steps.get("clf")
        all_classes = clf.classes_

    # 2) Softmax ile olasılık hesapla
    scores = np.asarray(scores)
    scores = scores - scores.max(axis=1, keepdims=True)
    exp_scores = np.exp(scores)
    probs = exp_scores / exp_scores.sum(axis=1, keepdims=True)

    # 3) Top-k indekslerini bul (yüksekten düşüğe sırala)
    k = int(max(1, min(k, probs.shape[1])))
    topk_idx = np.argsort(probs, axis=1)[:, -k:][:, ::-1]
    # 4) Etiket ve olasılıkları çıkar
    rows = probs.shape[0]
    topk_labels = np.empty((rows, k), dtype=object)
    topk_probs = np.empty((rows, k), dtype=float)
    for i in range(rows):
        idxs = topk_idx[i]
        topk_labels[i, :] = all_classes[idxs]
        topk_probs[i, :] = probs[i, idxs]
    return topk_labels, topk_probs, all_classes


def main():
    """
    Ana fonksiyon - Tüm süreci yönetir
    
    Bu fonksiyon şu adımları gerçekleştirir:
    1. Excel dosyasını okur ve başlık tespiti yapar
    2. Gerekli sütunları bulur ve veriyi temizler
    3. Etiketli ve etiketsiz verileri ayırır
    4. Model eğitimi ve hiperparametre optimizasyonu yapar
    5. Etiketsiz verilerin kategorilerini tahmin eder
    6. Sonuçları Excel dosyasına kaydeder
    7. Görselleştirmeler oluşturur
    
    Girdi: musteri sikayetleri.xlsx
    Çıktı: musteri_sikayetleri_tahminli.xlsx + görselleştirmeler
    """
    warnings.filterwarnings("ignore")

    # Dosya yollarını belirle
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(base_dir, "musteri sikayetleri.xlsx")

    # Girdi dosyasının varlığını kontrol et
    if not os.path.exists(input_path):
        print(f"Excel dosyası bulunamadı: {input_path}", file=sys.stderr)
        sys.exit(1)

    print("Excel okunuyor ve başlık tespiti yapılıyor...", flush=True)
    df = load_excel_with_header_detection(input_path)
    df = normalize_column_names(df)

    try:
        tarih_col, yorum_col, kategori_col = find_required_columns(df)
    except KeyError as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)

    # Ensure text column is string + clean Turkish text
    def clean_text_tr(text: str) -> str:
        if pd.isna(text):
            return ""
        t = str(text).strip().lower()
        # remove urls and emails
        t = re.sub(r"https?://\S+|www\.\S+", " ", t)
        t = re.sub(r"\b[\w\.-]+@[\w\.-]+\.[a-zA-Z]{2,}\b", " ", t)
        # collapse repeated characters (3+ -> 2)
        t = re.sub(r"(.)\1{2,}", r"\1\1", t)
        # collapse whitespace
        t = re.sub(r"\s+", " ", t)
        return t

    df[yorum_col] = df[yorum_col].astype(str).map(clean_text_tr)

    # Identify labeled vs unlabeled rows
    labeled_mask = df[kategori_col].notna() & (df[kategori_col].astype(str).str.strip() != "")
    unlabeled_mask = ~labeled_mask

    labeled_df = df[labeled_mask].copy()
    unlabeled_df = df[unlabeled_mask].copy()

    if labeled_df.empty:
        print("Uyarı: Eğitim için etiketli veri bulunamadı. Çıkılıyor.", file=sys.stderr)
        sys.exit(1)

    print(f"Etiketli satır sayısı: {len(labeled_df)}")
    print(f"Etiketsiz satır sayısı: {len(unlabeled_df)}")

    X_all = labeled_df[yorum_col]
    y_all = labeled_df[kategori_col].astype(str).str.strip()

    # Holdout evaluation if we have enough samples and at least 2 samples per class for stratification
    class_counts = y_all.value_counts()
    eval_mask = y_all.map(class_counts) >= 2
    X_eval = X_all[eval_mask]
    y_eval = y_all[eval_mask]
    do_holdout = len(X_eval) >= 50 and y_eval.nunique() > 1
    validation_accuracy: Optional[float] = None
    if do_holdout:
        X_train, X_val, y_train, y_val = train_test_split(
            X_eval, y_eval, test_size=0.2, random_state=42, stratify=y_eval
        )
        print("Hiperparametre optimizasyonu yapılıyor (küçük bir arama ızgarası)...")
        if VECTORIZER_MODE == 'tfidf' and PREFER_CHAR_ONLY:
            print("Char-only TF-IDF deneniyor...")
            best_model = tune_tfidf_with_analyzer(X_train, y_train, analyzer="char")
        elif VECTORIZER_MODE == 'tfidf' and ENSEMBLE_TFIDF:
            # Train two models: word-based and char-based, then ensemble
            print("Ensemble TF-IDF (word + char) deneniyor...")
            best_word = tune_tfidf_with_analyzer(X_train, y_train, analyzer="word")
            best_char = tune_tfidf_with_analyzer(X_train, y_train, analyzer="char")
            # Evaluate ensemble on validation to choose or keep both
            # Build class union
            classes_union = np.unique(np.concatenate([best_word.named_steps['clf'].classes_, best_char.named_steps['clf'].classes_]))
            sw = decision_scores(best_word, X_val, classes_union)
            sc = decision_scores(best_char, X_val, classes_union)
            s_ens = sw + sc
            y_pred = classes_union[s_ens.argmax(axis=1)]
            acc_ens = accuracy_score(y_val, y_pred)
            # Also compute single-model accuracies
            y_pred_w = best_word.predict(X_val)
            y_pred_c = best_char.predict(X_val)
            acc_w = accuracy_score(y_val, y_pred_w)
            acc_c = accuracy_score(y_val, y_pred_c)
            print(f"Val Acc (word): {acc_w:.4f} | (char): {acc_c:.4f} | (ensemble): {acc_ens:.4f}")
            # Keep both for final training; we'll ensemble at inference too
            best_model = (best_word, best_char)
        else:
            best_model = tune_hyperparameters(X_train, y_train)
        print("Doğrulama setinde değerlendirme yapılıyor...")
        if isinstance(best_model, tuple):
            best_word, best_char = best_model
            classes_union = np.unique(np.concatenate([best_word.named_steps['clf'].classes_, best_char.named_steps['clf'].classes_]))
            sw = decision_scores(best_word, X_val, classes_union)
            sc = decision_scores(best_char, X_val, classes_union)
            s_ens = sw + sc
            y_pred = classes_union[s_ens.argmax(axis=1)]
            acc = accuracy_score(y_val, y_pred)
        elif VECTORIZER_MODE == 'tfidf' and PREFER_CHAR_ONLY:
            y_pred = best_model.predict(X_val)
            acc = accuracy_score(y_val, y_pred)
        else:
            y_pred = best_model.predict(X_val)
            acc = accuracy_score(y_val, y_pred)
        validation_accuracy = float(acc)
        print(f"Doğrulama Accuracy: {acc:.4f}")
        print("Sınıflandırma Raporu:")
        report_txt = classification_report(y_val, y_pred)
        print(report_txt)
        report_dict = classification_report(y_val, y_pred, output_dict=True)

        # Kaydet: confusion matrix ve metrikler (isteğe bağlı)
        base_dir = os.path.dirname(os.path.abspath(__file__))
        if GENERATE_DIAGNOSTIC_PLOTS:
            cm = confusion_matrix(y_val, y_pred, labels=sorted(y_eval.unique()))
            plt.figure(figsize=(14, 10))
            sns.heatmap(cm, annot=False, fmt="d", cmap="Blues",
                        xticklabels=sorted(y_eval.unique()), yticklabels=sorted(y_eval.unique()))
            plt.title("Confusion Matrix (Validation)")
            plt.xlabel("Predicted")
            plt.ylabel("True")
            plt.tight_layout()
            plt.savefig(os.path.join(base_dir, "confusion_matrix_validation.png"))
            plt.close()

            # Sınıf dağılım grafiği
            plt.figure(figsize=(12, 4))
            class_counts.sort_values(ascending=False).plot(kind="bar")
            plt.title("Sınıf Dağılımı (Tüm Etiketli Veride)")
            plt.ylabel("Adet")
            plt.tight_layout()
            plt.savefig(os.path.join(base_dir, "sinif_dagilimi.png"))
            plt.close()

            # Per-class F1 bar chart
            labels = []
            f1_scores = []
            for lbl, metrics in report_dict.items():
                if lbl in {"accuracy", "macro avg", "weighted avg"}:
                    continue
                labels.append(lbl)
                f1_scores.append(metrics.get("f1-score", 0.0))
            plt.figure(figsize=(12, 4))
            sns.barplot(x=labels, y=f1_scores)
            plt.xticks(rotation=45, ha="right")
            plt.ylabel("F1-score")
            plt.title("Per-Class F1 (Validation)")
            plt.tight_layout()
            plt.savefig(os.path.join(base_dir, "per_class_f1.png"))
            plt.close()

        # Raporu JSON olarak kaydet
        with open(os.path.join(base_dir, "model_metrics.json"), "w", encoding="utf-8") as f:
            json.dump({
                "validation_accuracy": validation_accuracy,
                "classification_report": report_dict,
            }, f, ensure_ascii=False, indent=2)

        # Tüm etiketli veri ile nihai modeli eğit (en iyi hyperparametrelerle)
        if isinstance(best_model, tuple):
            best_word, best_char = best_model
            final_word = best_word.fit(X_all, y_all)
            final_char = best_char.fit(X_all, y_all)
            final_model = (final_word, final_char)
        elif VECTORIZER_MODE == 'tfidf' and PREFER_CHAR_ONLY:
            final_model = best_model.fit(X_all, y_all)
        else:
            final_model = best_model.fit(X_all, y_all)
    else:
        print("Doğrulama ayrımı atlandı (veri az/seyrek sınıf). Tek adım eğitim yapılıyor...")
        if VECTORIZER_MODE == 'tfidf' and PREFER_CHAR_ONLY:
            best_char = tune_tfidf_with_analyzer(X_all, y_all, analyzer="char")
            final_model = best_char.fit(X_all, y_all)
        elif VECTORIZER_MODE == 'tfidf' and ENSEMBLE_TFIDF:
            best_word = tune_tfidf_with_analyzer(X_all, y_all, analyzer="word")
            best_char = tune_tfidf_with_analyzer(X_all, y_all, analyzer="char")
            final_model = (best_word.fit(X_all, y_all), best_char.fit(X_all, y_all))
        else:
            final_model = build_pipeline().fit(X_all, y_all)

    # Predict for unlabeled rows with confidence using softmax over decision_function
    if not unlabeled_df.empty:
        print("Etiketsiz satırlar için kategoriler tahmin ediliyor...")
        # Predict and record confidence
        pred_labels, pred_conf, _ = predict_with_confidence_and_overrides(final_model, unlabeled_df[yorum_col], enable_overrides=False)
        # Top-3 tahminleri hesapla (ilk 3 en olası sınıf ve olasılıkları)
        topk_labels, topk_probs, _ = predict_topk(final_model, unlabeled_df[yorum_col], k=3)
        df.loc[unlabeled_mask, kategori_col] = pred_labels
        # Bireysel tahmin güveni ve satır bazlı model accuracy kolonları istenmiyor
        # Top-k kolonlarını ekle (sadece 2. ve 3. tahmin)
        if topk_labels.shape[1] >= 2:
            df.loc[unlabeled_mask, "2.tahmin"] = topk_labels[:, 1]
        if topk_labels.shape[1] >= 3:
            df.loc[unlabeled_mask, "3.tahmin"] = topk_labels[:, 2]
        # validation_accuracy tabloya kolon olarak eklenmeyecek (ayrı sheet tek hücre)

        # Pseudo-labeling: add confident unlabeled to training and refit (once)
        if ENABLE_PSEUDO_LABELING:
            confident_idx = unlabeled_df.index[pred_conf >= PSEUDO_LABEL_THRESHOLD]
            if len(confident_idx) > 0:
                print(f"Pseudo-labeling: {len(confident_idx)} yüksek güvenli örnek eklenecek ve model yeniden eğitilecek...")
                X_aug = pd.concat([X_all, df.loc[confident_idx, yorum_col]])
                y_aug = pd.concat([y_all, pd.Series(pred_labels[pred_conf >= PSEUDO_LABEL_THRESHOLD], index=confident_idx)])
                if VECTORIZER_MODE == 'tfidf' and ENSEMBLE_TFIDF:
                    # re-tune quickly on augmented data using previously chosen analyzers
                    best_word = tune_tfidf_with_analyzer(X_aug, y_aug, analyzer="word")
                    best_char = tune_tfidf_with_analyzer(X_aug, y_aug, analyzer="char")
                    final_model = (best_word.fit(X_aug, y_aug), best_char.fit(X_aug, y_aug))
                else:
                    final_model = build_pipeline().fit(X_aug, y_aug)
                # Re-predict unlabeled set after refit
                pred_labels2, pred_conf2, _ = predict_with_confidence_and_overrides(final_model, unlabeled_df[yorum_col], enable_overrides=False)
                # Top-3 tahminleri yeniden hesapla
                topk_labels2, topk_probs2, _ = predict_topk(final_model, unlabeled_df[yorum_col], k=3)
                df.loc[unlabeled_mask, kategori_col] = pred_labels2
                # bireysel güven kolonunu yazmıyoruz
                if topk_labels2.shape[1] >= 2:
                    df.loc[unlabeled_mask, "2.tahmin"] = topk_labels2[:, 1]
                if topk_labels2.shape[1] >= 3:
                    df.loc[unlabeled_mask, "3.tahmin"] = topk_labels2[:, 2]
                # model accuracy satır bazında yazılmayacak
    else:
        print("Etiketsiz satır yok; sadece model değerlendirmesi yapıldı.")

    # Save output Excel with only originally unlabeled rows and essential columns
    output_path = os.path.join(base_dir, "musteri sikayetleri_tahminli.xlsx")
    base_cols = [tarih_col, yorum_col, kategori_col]
    # Sadece 2. ve 3. tahmin kolonlarını ekle (varsa)
    topk_added_cols = [c for c in [
        "2.tahmin",
        "3.tahmin",
    ] if c in df.columns]
    cols_to_keep = base_cols + topk_added_cols
    out_df = df.loc[unlabeled_mask, cols_to_keep].copy()
    out_df.to_excel(output_path, index=False)
    print(f"Tamamlandı. Çıktı kaydedildi: {output_path}")

    # İsteğe bağlı: mevcut bir Excel dosyasının sağ tarafına Top-3 kolonlarını
    # biçimlendirmeyi (renk/sarı vurgular) bozmadan ekleyebiliriz. Böylece
    # kullanıcı tarafından sarı işaretlenen hücreler korunur.
    try:
        inplace_target = os.path.join(base_dir, "musteri sikayetleri_tahminli.xlsx")
        wb = load_workbook(inplace_target)
        ws = wb.active
        # Mevcut başlıkları oku ve yeni kolon başlıklarını ekle
        header_row = 1
        existing_headers = [cell.value for cell in ws[header_row]]
        new_headers = [
            "2.tahmin",
            "3.tahmin",
        ]
        # model confidence tek hücre: ana sayfanın en sağına yerleştir
        if validation_accuracy is not None:
            # varsa güncelle, yoksa ekle
            exist_mc = None
            for j, cell in enumerate(ws[header_row], start=1):
                val = str(cell.value) if cell.value is not None else ""
                if val.lower().startswith("model_confidence"):
                    exist_mc = j
                    break
            label = f"model_confidence ({validation_accuracy:.4f})"
            if exist_mc is None:
                ws.cell(row=header_row, column=ws.max_column + 1, value=label)
            else:
                ws.cell(row=header_row, column=exist_mc, value=label)
        # Eksik olanları sırayla ekle
        for h in new_headers:
            if h not in existing_headers and h in cols_to_keep:
                ws.cell(row=header_row, column=ws.max_column + 1, value=h)
        # Satır satır yaz (yalnızca unlabeled satırlar için)
        # DataFrame tarafında indeks eşleşmesini sağlamak için, out_df'i sözlük yapalım
        write_cols = [h for h in new_headers if h in cols_to_keep]
        col_name_to_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}
        # ws'de veri 2. satırdan başlar, out_df aynı sırayı taşır
        for i, (_, row) in enumerate(out_df.iterrows(), start=2):
            for cname in write_cols:
                target_col = col_name_to_index.get(cname)
                if target_col is None:
                    # başlık daha sonra eklenmiş olabilir; güncelle ve devam et
                    target_col = ws.max_column + 1
                    ws.cell(row=header_row, column=target_col, value=cname)
                    col_name_to_index[cname] = target_col
                ws.cell(row=i, column=target_col, value=row.get(cname))
        # Model accuracy tek hücre (ayrı sheet)
        if validation_accuracy is not None:
            if "model_info" in wb.sheetnames:
                del wb["model_info"]
            ws_info = wb.create_sheet("model_info")
            ws_info.cell(row=1, column=1, value="validation_accuracy")
            ws_info.cell(row=2, column=1, value=float(validation_accuracy))
        wb.save(inplace_target)
        print("Var olan tahminli dosyaya 2./3. tahmin kolonları eklendi (biçim korunarak).")
    except Exception as e:
        print(f"Top-3 kolonlarını yerinde ekleme atlandı/başarısız: {e}")

    # WordCloud görselleri (sadece overall varsayılan)
    try:
        if GENERATE_OVERALL_WORDCLOUD:
            print("WordCloud (overall) oluşturuluyor...")
            all_text = " ".join((df[yorum_col].astype(str).tolist()))
            wc = WordCloud(width=1200, height=800, background_color="white").generate(all_text)
            wc.to_file(os.path.join(base_dir, "wordcloud_overall.png"))

        if GENERATE_PER_CLASS_WORDCLOUDS:
            top_classes = class_counts.sort_values(ascending=False).head(6).index.tolist()
            for cls in top_classes:
                cls_text = " ".join((df[df[kategori_col].astype(str) == str(cls)][yorum_col].astype(str).tolist()))
                if not cls_text.strip():
                    continue
                wc_cls = WordCloud(width=1200, height=800, background_color="white").generate(cls_text)
                safe_cls = str(cls).replace("/", "-")
                wc_cls.to_file(os.path.join(base_dir, f"wordcloud_{safe_cls}.png"))
    except Exception as e:
        print(f"WordCloud oluşturulamadı: {e}")


if __name__ == "__main__":
    main()


